# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

import ast
import io
import re
from datetime import date, datetime

from odoo import Command, api, fields, models, _
from odoo.exceptions import ValidationError


class SurveySurvey(models.Model):
    _inherit = 'survey.survey'

    _IMPORT_ANSWER_HEADER_REGEX = re.compile(r'^(answer|correct|score)_?(\d+)$')

    slide_ids = fields.One2many(
        'slide.slide', 'survey_id', string="Certification Slides",
        help="The slides this survey is linked to through the e-learning application")
    slide_channel_ids = fields.One2many(
        'slide.channel', string="Certification Courses", compute='_compute_slide_channel_data',
        help="The courses this survey is linked to through the e-learning application",
        groups='website_slides.group_website_slides_officer')
    slide_channel_count = fields.Integer("Courses Count", compute='_compute_slide_channel_data', groups='website_slides.group_website_slides_officer')

    @api.depends('slide_ids.channel_id')
    def _compute_slide_channel_data(self):
        for survey in self:
            survey.slide_channel_ids = survey.slide_ids.mapped('channel_id')
            survey.slide_channel_count = len(survey.slide_channel_ids)

    @api.ondelete(at_uninstall=False)
    def _unlink_except_linked_to_course(self):
        # we consider it's ok to show certification names for people trying to delete courses
        # even if they don't have access to those surveys hence the sudo usage
        certifications = self.sudo().slide_ids.filtered(lambda slide: slide.slide_type == "certification").mapped('survey_id').exists()
        if certifications:
            certifications_course_mapping = [
                self.env._(
                    "- %(certification)s (Courses - %(courses)s)",
                    certification=certi.title,
                    courses=certi.slide_channel_ids.mapped("name"),
                )
                for certi in certifications
            ]
            raise ValidationError(_(
                'Uh-oh! You can’t delete surveys used as a Course Certification! Otherwise, students might think diplomas just grow on trees.\n'
                'The courses that need them are:\n%s',
                '\n'.join(certifications_course_mapping)))

    # ---------------------------------------------------------
    # Actions
    # ---------------------------------------------------------

    def action_survey_view_slide_channels(self):
        """ Redirect to the channels using the survey as a certification. Open
        in no-create as link between those two comes through a slide, hard to
        keep as default values. """
        action = self.env["ir.actions.actions"]._for_xml_id("website_slides.slide_channel_action_overview")
        action['display_name'] = _("Courses")
        if self.slide_channel_count == 1:
            action.update({'views': [(False, 'form')],
                           'res_id': self.slide_channel_ids[0].id})
        else:
            action.update({'views': [[False, 'list'], [False, 'form']],
                           'domain': [('id', 'in', self.slide_channel_ids.ids)]})
        action['context'] = dict(
            ast.literal_eval(action.get('context') or '{}'),  # sufficient in most cases
            create=False
        )
        return action

    # ---------------------------------------------------------
    # Business
    # ---------------------------------------------------------

    def _prepare_challenge_category(self):
        slide_survey = self.env['slide.slide'].search([('survey_id', '=', self.id)])
        return 'slides' if slide_survey else 'certification'

    def action_open_question_import_wizard(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Import Certification Questions'),
            'res_model': 'survey.question.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_survey_id': self.id,
            },
        }

    def action_download_question_import_template(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': '/slides_survey/certification/import_template',
            'target': 'self',
        }

    @api.model
    def _normalize_import_header(self, value):
        value = str(value or '').strip().lower()
        value = re.sub(r'[^a-z0-9]+', '_', value)
        return value.strip('_')

    @api.model
    def _format_import_cell(self, value):
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @api.model
    def _parse_import_boolean(self, value, row_index, column_name):
        if value in (None, ''):
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)
        normalized = str(value).strip().lower()
        if normalized in {'1', 'true', 'yes', 'y', 'x', 'correct'}:
            return True
        if normalized in {'0', 'false', 'no', 'n', 'wrong', 'incorrect'}:
            return False
        raise ValidationError(_(
            "Invalid value '%(value)s' in row %(row)s column '%(column)s'. Use TRUE/FALSE, YES/NO, or 1/0.",
            value=value,
            row=row_index,
            column=column_name,
        ))

    @api.model
    def _parse_import_float(self, value, row_index, column_name, default=None):
        if value in (None, ''):
            return default
        try:
            return float(value)
        except (TypeError, ValueError) as exc:
            raise ValidationError(_(
                "Invalid number '%(value)s' in row %(row)s column '%(column)s'.",
                value=value,
                row=row_index,
                column=column_name,
            )) from exc

    @api.model
    def _parse_import_int(self, value, row_index, column_name, default=None):
        if value in (None, ''):
            return default
        try:
            return int(float(value))
        except (TypeError, ValueError) as exc:
            raise ValidationError(_(
                "Invalid integer '%(value)s' in row %(row)s column '%(column)s'.",
                value=value,
                row=row_index,
                column=column_name,
            )) from exc

    @api.model
    def _parse_import_date(self, value, row_index, column_name):
        if value in (None, ''):
            return False
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        parsed = fields.Date.to_date(value)
        if not parsed:
            raise ValidationError(_(
                "Invalid date '%(value)s' in row %(row)s column '%(column)s'. Use YYYY-MM-DD.",
                value=value,
                row=row_index,
                column=column_name,
            ))
        return parsed

    @api.model
    def _parse_import_datetime(self, value, row_index, column_name):
        if value in (None, ''):
            return False
        if isinstance(value, datetime):
            return value
        parsed = fields.Datetime.to_datetime(value)
        if not parsed:
            raise ValidationError(_(
                "Invalid datetime '%(value)s' in row %(row)s column '%(column)s'. Use YYYY-MM-DD HH:MM:SS.",
                value=value,
                row=row_index,
                column=column_name,
            ))
        return parsed

    def _prepare_certification_questions_from_xlsx(self, file_content):
        self.ensure_one()
        try:
            from openpyxl import load_workbook  # noqa: PLC0415
        except ImportError as exc:
            raise ValidationError(_("The Python package 'openpyxl' is required to import certification questions from Excel.")) from exc

        try:
            workbook = load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
        except Exception as exc:
            raise ValidationError(_("The uploaded file could not be read. Please upload a valid .xlsx file.")) from exc

        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError(_("The uploaded Excel file is empty."))

        headers = [self._normalize_import_header(cell) for cell in rows[0]]
        if not any(headers):
            raise ValidationError(_("The first row of the Excel file must contain column headers."))

        required_header = 'title'
        if required_header not in headers:
            raise ValidationError(_("The Excel file must contain a 'Title' column."))

        header_index = {header: idx for idx, header in enumerate(headers) if header}
        answer_columns = {}
        for header, idx in header_index.items():
            match = self._IMPORT_ANSWER_HEADER_REGEX.match(header)
            if match:
                kind, number = match.groups()
                answer_columns.setdefault(int(number), {})[kind] = idx

        prepared_rows = []
        next_sequence = max(self.question_and_page_ids.mapped('sequence') or [0]) + 10
        current_section = None

        for row_index, row in enumerate(rows[1:], start=2):
            if not any(cell not in (None, '') for cell in row):
                continue

            title = self._format_import_cell(row[header_index['title']] if header_index['title'] < len(row) else None)
            if not title:
                raise ValidationError(_("Row %(row)s is missing the title.", row=row_index))

            question_type = self._normalize_import_header(
                row[header_index['question_type']] if 'question_type' in header_index and header_index['question_type'] < len(row) else 'simple_choice'
            ) or 'simple_choice'
            if question_type == 'multiple':
                question_type = 'multiple_choice'
            elif question_type == 'single':
                question_type = 'simple_choice'
            elif question_type == 'section':
                question_type = 'section'

            if question_type not in {'section', 'simple_choice', 'multiple_choice', 'char_box', 'text_box', 'numerical_box', 'date', 'datetime', 'scale'}:
                raise ValidationError(_(
                    "Unsupported question type '%(question_type)s' in row %(row)s.",
                    question_type=question_type,
                    row=row_index,
                ))

            section_name = self._format_import_cell(
                row[header_index['section']] if 'section' in header_index and header_index['section'] < len(row) else None
            )
            if section_name and section_name != current_section:
                prepared_rows.append({
                    'title': section_name,
                    'is_page': True,
                    'sequence': next_sequence,
                    'survey_id': self.id,
                })
                next_sequence += 10
                current_section = section_name

            if question_type == 'section':
                prepared_rows.append({
                    'title': title,
                    'is_page': True,
                    'sequence': next_sequence,
                    'survey_id': self.id,
                })
                next_sequence += 10
                current_section = title
                continue

            question_vals = {
                'survey_id': self.id,
                'title': title,
                'description': self._format_import_cell(
                    row[header_index['description']] if 'description' in header_index and header_index['description'] < len(row) else None
                ) or False,
                'question_type': question_type,
                'sequence': next_sequence,
                'constr_mandatory': self._parse_import_boolean(
                    row[header_index['mandatory']] if 'mandatory' in header_index and header_index['mandatory'] < len(row) else True,
                    row_index,
                    'Mandatory',
                ),
            }
            next_sequence += 10

            if question_type in {'simple_choice', 'multiple_choice'}:
                answer_values_list = []
                for answer_number in sorted(answer_columns):
                    column_map = answer_columns[answer_number]
                    answer_index = column_map.get('answer')
                    if answer_index is None or answer_index >= len(row):
                        continue
                    answer_text = self._format_import_cell(row[answer_index])
                    if not answer_text:
                        continue
                    is_correct = self._parse_import_boolean(
                        row[column_map['correct']] if column_map.get('correct') is not None and column_map['correct'] < len(row) else None,
                        row_index,
                        f'Correct {answer_number}',
                    )
                    answer_score = self._parse_import_float(
                        row[column_map['score']] if column_map.get('score') is not None and column_map['score'] < len(row) else None,
                        row_index,
                        f'Score {answer_number}',
                        default=(1.0 if is_correct else 0.0),
                    )
                    answer_values_list.append({
                        'sequence': len(answer_values_list) + 1,
                        'value': answer_text,
                        'is_correct': is_correct,
                        'answer_score': answer_score,
                    })

                if len(answer_values_list) < 2:
                    raise ValidationError(_("Row %(row)s must contain at least two answers.", row=row_index))
                if not any(answer['is_correct'] for answer in answer_values_list) or all(answer['is_correct'] for answer in answer_values_list):
                    raise ValidationError(_("Row %(row)s must contain at least one correct answer and one incorrect answer.", row=row_index))

                question_vals.update({
                    'is_scored_question': True,
                    'suggested_answer_ids': [Command.create(answer_values) for answer_values in answer_values_list],
                })
            elif question_type in {'char_box', 'text_box'}:
                question_vals.update({
                    'question_placeholder': self._format_import_cell(
                        row[header_index['placeholder']] if 'placeholder' in header_index and header_index['placeholder'] < len(row) else None
                    ) or False,
                })
                if question_type == 'char_box':
                    question_vals['validation_email'] = self._parse_import_boolean(
                        row[header_index['validation_email']] if 'validation_email' in header_index and header_index['validation_email'] < len(row) else None,
                        row_index,
                        'Validation Email',
                    )
                question_vals['is_scored_question'] = False
            elif question_type == 'numerical_box':
                question_vals.update({
                    'is_scored_question': True,
                    'answer_numerical_box': self._parse_import_float(
                        row[header_index['correct_value']] if 'correct_value' in header_index and header_index['correct_value'] < len(row) else None,
                        row_index,
                        'Correct Value',
                    ),
                    'answer_score': self._parse_import_float(
                        row[header_index['score']] if 'score' in header_index and header_index['score'] < len(row) else None,
                        row_index,
                        'Score',
                        default=1.0,
                    ),
                })
            elif question_type == 'date':
                question_vals.update({
                    'is_scored_question': True,
                    'answer_date': self._parse_import_date(
                        row[header_index['correct_value']] if 'correct_value' in header_index and header_index['correct_value'] < len(row) else None,
                        row_index,
                        'Correct Value',
                    ),
                    'answer_score': self._parse_import_float(
                        row[header_index['score']] if 'score' in header_index and header_index['score'] < len(row) else None,
                        row_index,
                        'Score',
                        default=1.0,
                    ),
                })
            elif question_type == 'datetime':
                question_vals.update({
                    'is_scored_question': True,
                    'answer_datetime': self._parse_import_datetime(
                        row[header_index['correct_value']] if 'correct_value' in header_index and header_index['correct_value'] < len(row) else None,
                        row_index,
                        'Correct Value',
                    ),
                    'answer_score': self._parse_import_float(
                        row[header_index['score']] if 'score' in header_index and header_index['score'] < len(row) else None,
                        row_index,
                        'Score',
                        default=1.0,
                    ),
                })
            elif question_type == 'scale':
                scale_min = self._parse_import_int(
                    row[header_index['scale_min']] if 'scale_min' in header_index and header_index['scale_min'] < len(row) else 0,
                    row_index,
                    'Scale Min',
                    default=0,
                )
                scale_max = self._parse_import_int(
                    row[header_index['scale_max']] if 'scale_max' in header_index and header_index['scale_max'] < len(row) else 10,
                    row_index,
                    'Scale Max',
                    default=10,
                )
                question_vals.update({
                    'is_scored_question': False,
                    'scale_min': scale_min,
                    'scale_max': scale_max,
                    'scale_min_label': self._format_import_cell(row[header_index['scale_min_label']]) if 'scale_min_label' in header_index and header_index['scale_min_label'] < len(row) else False,
                    'scale_mid_label': self._format_import_cell(row[header_index['scale_mid_label']]) if 'scale_mid_label' in header_index and header_index['scale_mid_label'] < len(row) else False,
                    'scale_max_label': self._format_import_cell(row[header_index['scale_max_label']]) if 'scale_max_label' in header_index and header_index['scale_max_label'] < len(row) else False,
                })

            prepared_rows.append(question_vals)

        if not prepared_rows:
            raise ValidationError(_("The Excel file does not contain any rows to import."))
        return prepared_rows

    def import_certification_questions_from_xlsx(self, file_content, replace_existing=False):
        self.ensure_one()
        prepared_rows = self._prepare_certification_questions_from_xlsx(file_content)
        if replace_existing:
            self.question_and_page_ids.unlink()
        created_records = self.env['survey.question']

        for vals in prepared_rows:
            question = self.env['survey.question'].create(vals)
            created_records |= question
        return created_records
