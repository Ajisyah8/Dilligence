# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.

import io
import re

from odoo import api, fields, models, _
from odoo.exceptions import ValidationError


class SlideQuestion(models.Model):
    _name = 'slide.question'
    _rec_name = "question"
    _description = "Content Quiz Question"
    _order = "sequence"

    sequence = fields.Integer("Sequence")
    question = fields.Char("Question Name", required=True, translate=True)
    slide_id = fields.Many2one('slide.slide', string="Content", required=True, index=True, ondelete='cascade')
    answer_ids = fields.One2many('slide.answer', 'question_id', string="Answer", copy=True)
    answers_validation_error = fields.Char("Error on Answers", compute='_compute_answers_validation_error')
    # statistics
    attempts_count = fields.Integer(compute='_compute_statistics', groups='website_slides.group_website_slides_officer')
    attempts_avg = fields.Float(compute="_compute_statistics", digits=(6, 2), groups='website_slides.group_website_slides_officer')
    done_count = fields.Integer(compute="_compute_statistics", groups='website_slides.group_website_slides_officer')

    _IMPORT_HEADER_REGEX = re.compile(r'^(answer|correct|comment)_?(\d+)$')

    @api.constrains('answer_ids')
    def _check_answers_integrity(self):
        questions_to_fix = [
            f'- {question.slide_id.name}: {question.question}'
            for question in self
            if question.answers_validation_error
        ]
        if questions_to_fix:
            raise ValidationError(_(
                'All questions must have at least one correct answer and one incorrect answer: \n%s\n',
                '\n'.join(questions_to_fix)))

    @api.depends('slide_id')
    def _compute_statistics(self):
        slide_partners = self.env['slide.slide.partner'].sudo().search([('slide_id', 'in', self.slide_id.ids)])
        slide_stats = dict((s.slide_id.id, dict({'attempts_count': 0, 'attempts_unique': 0, 'done_count': 0})) for s in slide_partners)

        for slide_partner in slide_partners:
            slide_stats[slide_partner.slide_id.id]['attempts_count'] += slide_partner.quiz_attempts_count
            slide_stats[slide_partner.slide_id.id]['attempts_unique'] += 1
            if slide_partner.completed:
                slide_stats[slide_partner.slide_id.id]['done_count'] += 1

        for question in self:
            stats = slide_stats.get(question.slide_id.id)
            question.attempts_count = stats.get('attempts_count', 0) if stats else 0
            question.attempts_avg = stats.get('attempts_count', 0) / stats.get('attempts_unique', 1) if stats else 0
            question.done_count = stats.get('done_count', 0) if stats else 0

    @api.depends('answer_ids', 'answer_ids.is_correct')
    def _compute_answers_validation_error(self):
        for question in self:
            correct = question.answer_ids.filtered('is_correct')
            question.answers_validation_error = _(
                'This question must have at least one correct answer and one incorrect answer.'
            ) if not correct or correct == question.answer_ids else ''

    @api.model
    def _normalize_import_header(self, value):
        value = str(value or '').strip().lower()
        value = re.sub(r'[^a-z0-9]+', '_', value)
        return value.strip('_')

    @api.model
    def _format_import_cell(self, value):
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @api.model
    def _parse_import_boolean(self, value, row_index, column_name):
        if value in (None, ''):
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return bool(value)

        normalized = str(value).strip().lower()
        if normalized in {'1', 'true', 'yes', 'y', 'x', 'correct'}:
            return True
        if normalized in {'0', 'false', 'no', 'n', 'wrong', 'incorrect'}:
            return False
        raise ValidationError(_(
            "Invalid value '%(value)s' in row %(row)s column '%(column)s'. Use TRUE/FALSE, YES/NO, or 1/0.",
            value=value,
            row=row_index,
            column=column_name,
        ))

    @api.model
    def _prepare_questions_from_xlsx(self, slide, file_content):
        try:
            from openpyxl import load_workbook  # noqa: PLC0415
        except ImportError as exc:
            raise ValidationError(_("The Python package 'openpyxl' is required to import quiz questions from Excel.")) from exc

        try:
            workbook = load_workbook(io.BytesIO(file_content), data_only=True, read_only=True)
        except Exception as exc:
            raise ValidationError(_("The uploaded file could not be read. Please upload a valid .xlsx file.")) from exc

        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError(_("The uploaded Excel file is empty."))

        headers = [self._normalize_import_header(cell) for cell in rows[0]]
        if not any(headers):
            raise ValidationError(_("The first row of the Excel file must contain column headers."))

        if 'question' not in headers:
            raise ValidationError(_("The Excel file must contain a 'Question' column."))

        answer_columns = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            match = self._IMPORT_HEADER_REGEX.match(header)
            if not match:
                continue
            column_type, column_number = match.groups()
            answer_columns.setdefault(int(column_number), {})[column_type] = index

        if not answer_columns:
            raise ValidationError(_("The Excel file must contain at least one 'Answer 1' column."))

        prepared_questions = []
        question_column_index = headers.index('question')

        for row_index, row in enumerate(rows[1:], start=2):
            if not any(cell not in (None, '') for cell in row):
                continue

            question_title = self._format_import_cell(row[question_column_index] if question_column_index < len(row) else None)
            if not question_title:
                raise ValidationError(_("Row %(row)s is missing the question title.", row=row_index))

            answers = []
            for answer_number in sorted(answer_columns):
                column_map = answer_columns[answer_number]
                answer_index = column_map.get('answer')
                if answer_index is None:
                    continue

                answer_text = self._format_import_cell(row[answer_index] if answer_index < len(row) else None)
                if not answer_text:
                    continue

                correct_index = column_map.get('correct')
                comment_index = column_map.get('comment')
                answers.append({
                    'sequence': len(answers) + 1,
                    'text_value': answer_text,
                    'is_correct': self._parse_import_boolean(
                        row[correct_index] if correct_index is not None and correct_index < len(row) else None,
                        row_index,
                        f'Correct {answer_number}',
                    ),
                    'comment': self._format_import_cell(
                        row[comment_index] if comment_index is not None and comment_index < len(row) else None
                    ),
                })

            if len(answers) < 2:
                raise ValidationError(_(
                    "Row %(row)s must contain at least two answers.",
                    row=row_index,
                ))

            if all(answer['is_correct'] for answer in answers) or not any(answer['is_correct'] for answer in answers):
                raise ValidationError(_(
                    "Row %(row)s must contain at least one correct answer and one incorrect answer.",
                    row=row_index,
                ))

            prepared_questions.append({
                'slide_id': slide.id,
                'question': question_title,
                'answer_ids': answers,
            })

        if not prepared_questions:
            raise ValidationError(_("The Excel file does not contain any question rows to import."))

        return prepared_questions


class SlideAnswer(models.Model):
    _name = 'slide.answer'
    _rec_name = "text_value"
    _description = "Slide Question's Answer"
    _order = 'question_id, sequence, id'

    sequence = fields.Integer("Sequence")
    question_id = fields.Many2one('slide.question', string="Question", required=True, index=True, ondelete='cascade')
    text_value = fields.Char("Answer", required=True, translate=True)
    is_correct = fields.Boolean("Is correct answer")
    comment = fields.Text("Comment", translate=True, help='This comment will be displayed to the user if they select this answer')
